import requests
from lxml import html
from urllib.parse import urljoin
import os
import re
from tqdm import tqdm
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows

# Регулярні вирази для шаблонів URL другого і третього рівнів
level2_pattern = re.compile(r"https://data\.loda\.gov\.ua/dataset/[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}")
level3_pattern = re.compile(r"https://data\.loda\.gov\.ua/dataset/[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}/resource/[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}$")

# Функція для перевірки наявності елементу на сторінці
def check_element(url, xpath):
    response = requests.get(url)
    if response.status_code == 200:
        tree = html.fromstring(response.content)
        element = tree.xpath(xpath)
        if element:
            return True
    return False

# Функція для отримання всіх посилань на сторінці
def get_links(url):
    response = requests.get(url)
    if response.status_code == 200:
        tree = html.fromstring(response.content)
        links = tree.xpath('//a[@href]')
        return [urljoin(url, link.get('href')) for link in links]
    return []

# Функція для отримання тексту за xpath
def get_text(url, xpath):
    response = requests.get(url)
    if response.status_code == 200:
        tree = html.fromstring(response.content)
        element = tree.xpath(xpath)
        if element:
            return element[0].text_content().strip()
    return ""

# Початкова сторінка та xpath елементу для перевірки
base_url = "https://data.loda.gov.ua/dataset/"
xpath_check = "/html/body/div[2]/div/div[2]/section/div/div[1]/ul/li[2]"

# Xpath для додаткової інформації
xpath_level2_title = "/html/body/div[2]/div/div[2]/div/article/div/h1"
xpath_level2_aside = "//th[contains(text(), 'Відповідальна особа')]/following-sibling::td"
xpath_level2_info = "//section[contains(@class, 'module-content')]//h1[contains(@class, 'heading')]"
xpath_level3_title = "/html/body/div[2]/div/div[2]/section/div/h1"

# Список для зберігання результатів
results = []

# Збір сторінок другого рівня з усіх підсторінок головної сторінки
level2_urls = set()
for page_num in range(1, 11):
    page_url = f"{base_url}?page={page_num}"
    level2_urls.update(url for url in get_links(page_url) if level2_pattern.match(url))

# Збір сторінок третього рівня та перевірка наявності елементу
processed_level3_urls = set()

for url in tqdm(level2_urls, desc="Processing URLs", unit="URL"):
    level3_urls = [url for url in get_links(url) if level3_pattern.match(url)]
    
    # Отримання додаткової інформації для сторінок другого рівня
    level2_title = get_text(url, xpath_level2_title)
    level2_info = get_text(url, xpath_level2_info)
    level2_aside = get_text(url, xpath_level2_aside)

    for sub_url in level3_urls:
        if sub_url not in processed_level3_urls:
            processed_level3_urls.add(sub_url)
            level3_title = get_text(sub_url, xpath_level3_title)
            element_found = check_element(sub_url, xpath_check)
            results.append((url, level2_title, sub_url, level3_title, level2_info, level2_aside, 'Yes' if element_found else 'No'))

# Перетворення результатів у DataFrame
df = pd.DataFrame(results, columns=[
    "DatasetURL", "DatasetTitle", "ResourceURL", "ResourceTitle", 
    "Organization", "ResponsiblePerson", "ApiFound"
])

# Запис результатів у Excel-файл у тій же папці, де знаходиться скрипт
script_dir = os.path.dirname(os.path.realpath(__file__))
results_file_path = os.path.join(script_dir, 'results.xlsx')
df.to_excel(results_file_path, index=False)

# Завантаження створеного Excel-файлу
wb = load_workbook(results_file_path)
ws = wb.active

# Додавання умовного форматування
red_fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")

rule = FormulaRule(formula=['$G1="No"'], fill=red_fill)
ws.conditional_formatting.add(f"A1:G{ws.max_row}", rule)

# Збереження Excel-файлу з умовним форматуванням
wb.save(results_file_path)

print(f"Результати збережено у файл {results_file_path}")
